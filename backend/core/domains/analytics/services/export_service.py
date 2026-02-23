# backend/core/domains/analytics/services/export_service.py
import csv
import io
from datetime import datetime

from django.http import HttpResponse

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """Service for exporting analytics data to various formats."""

    @staticmethod
    def export_to_csv(data: list, filename: str, headers: list = None) -> HttpResponse:
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries to export
            filename: Name of the file (without extension)
            headers: Optional list of column headers (uses dict keys if not provided)
        """
        if not data:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
            return response

        output = io.StringIO()

        # Get headers from first row if not provided
        if not headers:
            headers = list(data[0].keys())

        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response

    @staticmethod
    def export_to_excel(data: list, filename: str, headers: list = None, sheet_name: str = "Data") -> HttpResponse:
        """
        Export data to Excel format.

        Args:
            data: List of dictionaries to export
            filename: Name of the file (without extension)
            headers: Optional list of column headers
            sheet_name: Name of the Excel sheet
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl not installed
            return ExportService.export_to_csv(data, filename, headers)

        if not data:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Get headers from first row if not provided
            if not headers:
                headers = list(data[0].keys())

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response

    @staticmethod
    def export_customers(data: list, format: str = "csv") -> HttpResponse:
        """Export customer list data."""
        filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        headers = ["id", "email", "full_name", "total_events", "completed_events", "total_spent", "created_at"]

        if format == "excel":
            return ExportService.export_to_excel(data, filename, headers, "Customers")
        return ExportService.export_to_csv(data, filename, headers)

    @staticmethod
    def export_bookings_summary(data: list, format: str = "csv") -> HttpResponse:
        """Export bookings summary data."""
        filename = f"bookings_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        headers = [
            "period",
            "total_bookings",
            "confirmed_bookings",
            "completed_bookings",
            "cancelled_bookings",
            "total_revenue",
        ]

        if format == "excel":
            return ExportService.export_to_excel(data, filename, headers, "Bookings Summary")
        return ExportService.export_to_csv(data, filename, headers)

    @staticmethod
    def export_revenue_report(data: list, format: str = "csv") -> HttpResponse:
        """Export revenue by type data."""
        filename = f"revenue_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        headers = ["name", "category", "booking_count", "total_revenue", "avg_revenue", "total_participants"]

        if format == "excel":
            return ExportService.export_to_excel(data, filename, headers, "Revenue Report")
        return ExportService.export_to_csv(data, filename, headers)

    @staticmethod
    def export_lead_sources(data: list, format: str = "csv") -> HttpResponse:
        """Export lead source report data."""
        filename = f"lead_sources_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        headers = ["label", "lead_count", "converted_count", "conversion_rate", "total_value"]

        if format == "excel":
            return ExportService.export_to_excel(data, filename, headers, "Lead Sources")
        return ExportService.export_to_csv(data, filename, headers)
