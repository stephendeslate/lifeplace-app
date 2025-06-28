# backend/core/domains/analytics/utils/export_helpers.py
import io
import csv
import json
import logging
import tempfile
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
from decimal import Decimal
import os

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Exception raised when export operations fail"""
    pass


class ReportExporter:
    """
    Utility class for exporting analytics reports in various formats
    """
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> io.StringIO:
        """Export data to CSV format"""
        try:
            output = io.StringIO()
            
            if not data:
                writer = csv.writer(output)
                writer.writerow(['No data available'])
                output.seek(0)
                return output
            
            # Get all unique keys for headers
            headers = set()
            for row in data:
                headers.update(row.keys())
            headers = sorted(list(headers))
            
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            
            for row in data:
                # Convert complex types to strings
                clean_row = {}
                for key, value in row.items():
                    if isinstance(value, (dict, list)):
                        clean_row[key] = json.dumps(value)
                    elif isinstance(value, Decimal):
                        clean_row[key] = str(value)
                    elif value is None:
                        clean_row[key] = ''
                    else:
                        clean_row[key] = str(value)
                writer.writerow(clean_row)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise ExportError(f"CSV export failed: {str(e)}")
    
    @staticmethod
    def export_to_excel(report_data: Dict[str, Any], filename: str = None) -> io.BytesIO:
        """Export report data to Excel format with charts and formatting"""
        if not OPENPYXL_AVAILABLE:
            raise ExportError("openpyxl is not available for Excel export")
        
        try:
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create summary sheet
            ReportExporter._create_excel_summary_sheet(workbook, report_data)
            
            # Create data sheets for each metric
            ReportExporter._create_excel_data_sheets(workbook, report_data)
            
            # Create charts sheet
            ReportExporter._create_excel_charts_sheet(workbook, report_data)
            
            workbook.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}")
    
    @staticmethod
    def export_to_pdf(report_data: Dict[str, Any], filename: str = None) -> io.BytesIO:
        """Export report data to PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise ExportError("reportlab is not available for PDF export")
        
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            report_name = report_data.get('report', {}).get('name', 'Analytics Report')
            story.append(Paragraph(report_name, title_style))
            story.append(Spacer(1, 20))
            
            # Add report metadata
            ReportExporter._add_pdf_metadata(story, report_data, styles)
            
            # Add metrics summary
            ReportExporter._add_pdf_metrics(story, report_data, styles)
            
            # Add charts if available
            ReportExporter._add_pdf_charts(story, report_data, styles)
            
            doc.build(story)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}")
            raise ExportError(f"PDF export failed: {str(e)}")
    
    @staticmethod
    def export_to_html(report_data: Dict[str, Any], template_name: str = None) -> str:
        """Export report data to HTML format"""
        try:
            template_name = template_name or 'analytics/reports/default_report.html'
            
            context = {
                'report_data': report_data,
                'generated_at': datetime.now(),
                'title': report_data.get('report', {}).get('name', 'Analytics Report')
            }
            
            html_content = render_to_string(template_name, context)
            return html_content
            
        except Exception as e:
            logger.error(f"Error exporting to HTML: {str(e)}")
            raise ExportError(f"HTML export failed: {str(e)}")
    
    @staticmethod
    def export_to_json(report_data: Dict[str, Any], indent: int = 2) -> str:
        """Export report data to JSON format"""
        try:
            # Convert Decimal objects to strings for JSON serialization
            def decimal_default(obj):
                if isinstance(obj, Decimal):
                    return str(obj)
                elif isinstance(obj, datetime):
                    return obj.isoformat()
                raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
            
            return json.dumps(report_data, indent=indent, default=decimal_default)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}")
            raise ExportError(f"JSON export failed: {str(e)}")
    
    # Excel helper methods
    @staticmethod
    def _create_excel_summary_sheet(workbook, report_data):
        """Create summary sheet in Excel workbook"""
        ws = workbook.create_sheet("Summary", 0)
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        report_name = report_data.get('report', {}).get('name', 'Analytics Report')
        ws['A1'] = report_name
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Report info
        row = 3
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        execution_params = report_data.get('execution_params', {})
        time_range = execution_params.get('time_range', {})
        if time_range:
            ws[f'A{row}'] = 'Date Range:'
            ws[f'B{row}'] = f"{time_range.get('start_date', 'N/A')} to {time_range.get('end_date', 'N/A')}"
            row += 1
        
        # Metrics summary
        row += 2
        ws[f'A{row}'] = 'Metrics Summary'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # Headers
        ws[f'A{row}'] = 'Metric Name'
        ws[f'B{row}'] = 'Value'
        ws[f'C{row}'] = 'Type'
        ws[f'D{row}'] = 'Format'
        
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        
        # Metric data
        result_data = report_data.get('result_data', {})
        for metric_name, metric_info in result_data.items():
            if isinstance(metric_info, dict) and 'value' in metric_info:
                ws[f'A{row}'] = metric_name
                ws[f'B{row}'] = metric_info.get('value', 'N/A')
                ws[f'C{row}'] = metric_info.get('metric_type', 'N/A')
                ws[f'D{row}'] = metric_info.get('display_format', 'N/A')
                row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_excel_data_sheets(workbook, report_data):
        """Create individual data sheets for each metric"""
        result_data = report_data.get('result_data', {})
        
        for metric_name, metric_info in result_data.items():
            if not isinstance(metric_info, dict):
                continue
                
            # Create sheet for this metric
            safe_name = metric_name[:31].replace('/', '_').replace('\\', '_')  # Excel sheet name limits
            ws = workbook.create_sheet(safe_name)
            
            # Header
            ws['A1'] = metric_name
            ws['A1'].font = Font(bold=True, size=14)
            
            # Metric details
            row = 3
            ws[f'A{row}'] = 'Value:'
            ws[f'B{row}'] = metric_info.get('value', 'N/A')
            row += 1
            
            ws[f'A{row}'] = 'Type:'
            ws[f'B{row}'] = metric_info.get('metric_type', 'N/A')
            row += 1
            
            ws[f'A{row}'] = 'Format:'
            ws[f'B{row}'] = metric_info.get('display_format', 'N/A')
            row += 1
            
            if 'error' in metric_info:
                ws[f'A{row}'] = 'Error:'
                ws[f'B{row}'] = metric_info['error']
                ws[f'B{row}'].font = Font(color="FF0000")
    
    @staticmethod
    def _create_excel_charts_sheet(workbook, report_data):
        """Create charts sheet in Excel workbook"""
        ws = workbook.create_sheet("Charts")
        
        ws['A1'] = 'Visual Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Note about charts
        ws['A3'] = 'Chart visualizations would be generated here based on historical data.'
        ws['A4'] = 'This requires additional time-series data collection implementation.'
    
    # PDF helper methods
    @staticmethod
    def _add_pdf_metadata(story, report_data, styles):
        """Add metadata section to PDF"""
        story.append(Paragraph("Report Information", styles['Heading2']))
        
        metadata = [
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Report Type:', report_data.get('report', {}).get('report_type', 'N/A')],
        ]
        
        execution_params = report_data.get('execution_params', {})
        time_range = execution_params.get('time_range', {})
        if time_range:
            metadata.append(['Date Range:', f"{time_range.get('start_date', 'N/A')} to {time_range.get('end_date', 'N/A')}"])
        
        table = Table(metadata)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    @staticmethod
    def _add_pdf_metrics(story, report_data, styles):
        """Add metrics section to PDF"""
        story.append(Paragraph("Metrics Summary", styles['Heading2']))
        
        result_data = report_data.get('result_data', {})
        if not result_data:
            story.append(Paragraph("No metrics data available.", styles['Normal']))
            return
        
        # Create metrics table
        table_data = [['Metric Name', 'Value', 'Type', 'Format']]
        
        for metric_name, metric_info in result_data.items():
            if isinstance(metric_info, dict) and 'value' in metric_info:
                table_data.append([
                    metric_name,
                    str(metric_info.get('value', 'N/A')),
                    metric_info.get('metric_type', 'N/A'),
                    metric_info.get('display_format', 'N/A')
                ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    @staticmethod
    def _add_pdf_charts(story, report_data, styles):
        """Add charts section to PDF"""
        story.append(Paragraph("Visual Analysis", styles['Heading2']))
        story.append(Paragraph("Chart visualizations would be generated here based on historical data.", styles['Normal']))
        story.append(Paragraph("This requires additional time-series data collection implementation.", styles['Normal']))


class DashboardExporter:
    """Export utility specifically for dashboards"""
    
    @staticmethod
    def export_dashboard_to_pdf(dashboard_data: Dict[str, Any]) -> io.BytesIO:
        """Export dashboard to PDF with widget visualizations"""
        if not REPORTLAB_AVAILABLE:
            raise ExportError("reportlab is not available for PDF export")
        
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Dashboard title
            dashboard = dashboard_data.get('dashboard', {})
            title = dashboard.get('name', 'Dashboard')
            
            title_style = ParagraphStyle(
                'DashboardTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=20,
                alignment=1
            )
            
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 20))
            
            # Dashboard info
            info_data = [
                ['Dashboard Type:', dashboard.get('dashboard_type', 'N/A')],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Auto Refresh:', f"{dashboard.get('auto_refresh_interval', 300)} seconds"],
            ]
            
            time_range = dashboard_data.get('time_range', {})
            if time_range:
                info_data.append(['Time Range:', time_range.get('label', 'N/A')])
            
            info_table = Table(info_data)
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Widgets section
            story.append(Paragraph("Dashboard Widgets", styles['Heading2']))
            
            widgets_data = dashboard_data.get('widgets_data', [])
            if not widgets_data:
                story.append(Paragraph("No widgets configured.", styles['Normal']))
            else:
                DashboardExporter._add_widgets_to_pdf(story, widgets_data, styles)
            
            doc.build(story)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting dashboard to PDF: {str(e)}")
            raise ExportError(f"Dashboard PDF export failed: {str(e)}")
    
    @staticmethod
    def _add_widgets_to_pdf(story, widgets_data, styles):
        """Add widgets information to PDF"""
        for widget_data in widgets_data:
            widget = widget_data.get('widget', {})
            value = widget_data.get('value')
            error = widget_data.get('error')
            
            # Widget title
            widget_title = widget.get('title', 'Unnamed Widget')
            story.append(Paragraph(widget_title, styles['Heading3']))
            
            # Widget details
            widget_info = [
                ['Type:', widget.get('widget_type', 'N/A')],
                ['Size:', widget.get('size', 'N/A')],
                ['Time Range:', widget.get('time_range', 'N/A')],
            ]
            
            if error:
                widget_info.append(['Error:', error])
                value_text = 'Error occurred'
            else:
                widget_info.append(['Value:', str(value) if value is not None else 'No data'])
                value_text = str(value) if value is not None else 'No data'
            
            widget_table = Table(widget_info)
            widget_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(widget_table)
            story.append(Spacer(1, 15))
    
    @staticmethod
    def export_dashboard_to_excel(dashboard_data: Dict[str, Any]) -> io.BytesIO:
        """Export dashboard to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise ExportError("openpyxl is not available for Excel export")
        
        try:
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Dashboard summary sheet
            ws = workbook.create_sheet("Dashboard Summary", 0)
            
            dashboard = dashboard_data.get('dashboard', {})
            
            # Title
            ws['A1'] = dashboard.get('name', 'Dashboard')
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')
            
            # Dashboard info
            row = 3
            ws[f'A{row}'] = 'Type:'
            ws[f'B{row}'] = dashboard.get('dashboard_type', 'N/A')
            row += 1
            
            ws[f'A{row}'] = 'Generated:'
            ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            row += 1
            
            time_range = dashboard_data.get('time_range', {})
            if time_range:
                ws[f'A{row}'] = 'Time Range:'
                ws[f'B{row}'] = time_range.get('label', 'N/A')
                row += 1
            
            # Widgets section
            row += 2
            ws[f'A{row}'] = 'Widgets Summary'
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            # Widget headers
            headers = ['Widget Name', 'Type', 'Value', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            # Widget data
            widgets_data = dashboard_data.get('widgets_data', [])
            for widget_data in widgets_data:
                widget = widget_data.get('widget', {})
                value = widget_data.get('value')
                error = widget_data.get('error')
                
                ws.cell(row=row, column=1).value = widget.get('title', 'Unnamed Widget')
                ws.cell(row=row, column=2).value = widget.get('widget_type', 'N/A')
                ws.cell(row=row, column=3).value = str(value) if value is not None else 'No data'
                ws.cell(row=row, column=4).value = 'Error' if error else 'OK'
                
                if error:
                    ws.cell(row=row, column=4).font = Font(color="FF0000")
                
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting dashboard to Excel: {str(e)}")
            raise ExportError(f"Dashboard Excel export failed: {str(e)}")


class DataExportService:
    """Service for handling various data export operations"""
    
    @staticmethod
    def export_analytics_data(data: Union[Dict, List], 
                            format_type: str, 
                            filename: str = None) -> Union[str, io.BytesIO, io.StringIO]:
        """
        Export analytics data in specified format
        
        Args:
            data: Data to export (dict for reports, list for raw data)
            format_type: Export format ('csv', 'excel', 'pdf', 'html', 'json')
            filename: Optional filename (without extension)
            
        Returns:
            Exported data as file-like object or string
        """
        try:
            if format_type.lower() == 'csv':
                if isinstance(data, dict):
                    # Convert report data to list format for CSV
                    csv_data = DataExportService._report_to_csv_format(data)
                else:
                    csv_data = data
                return ReportExporter.export_to_csv(csv_data, filename)
                
            elif format_type.lower() == 'excel':
                if isinstance(data, list):
                    # Convert list data to report format for Excel
                    report_data = {'result_data': {'Data': {'value': len(data)}}, 'raw_data': data}
                else:
                    report_data = data
                return ReportExporter.export_to_excel(report_data, filename)
                
            elif format_type.lower() == 'pdf':
                if isinstance(data, list):
                    # Convert list data to report format for PDF
                    report_data = {'result_data': {'Data': {'value': len(data)}}, 'raw_data': data}
                else:
                    report_data = data
                return ReportExporter.export_to_pdf(report_data, filename)
                
            elif format_type.lower() == 'html':
                if isinstance(data, list):
                    # Convert list data to report format for HTML
                    report_data = {'result_data': {'Data': {'value': len(data)}}, 'raw_data': data}
                else:
                    report_data = data
                return ReportExporter.export_to_html(report_data)
                
            elif format_type.lower() == 'json':
                return ReportExporter.export_to_json(data)
                
            else:
                raise ExportError(f"Unsupported export format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error in data export: {str(e)}")
            raise ExportError(f"Data export failed: {str(e)}")
    
    @staticmethod
    def _report_to_csv_format(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Convert report data to CSV-friendly format"""
        csv_data = []
        
        result_data = report_data.get('result_data', {})
        for metric_name, metric_info in result_data.items():
            if isinstance(metric_info, dict):
                csv_data.append({
                    'metric_name': metric_name,
                    'value': metric_info.get('value', ''),
                    'metric_type': metric_info.get('metric_type', ''),
                    'display_format': metric_info.get('display_format', ''),
                    'error': metric_info.get('error', '')
                })
        
        return csv_data if csv_data else [{'message': 'No data available'}]
    
    @staticmethod
    def create_download_response(file_data: Union[str, io.BytesIO, io.StringIO], 
                               filename: str, 
                               content_type: str) -> HttpResponse:
        """Create HTTP response for file download"""
        try:
            if isinstance(file_data, str):
                # String data (HTML, JSON)
                response = HttpResponse(file_data, content_type=content_type)
            elif isinstance(file_data, io.StringIO):
                # CSV data
                response = HttpResponse(file_data.getvalue(), content_type=content_type)
            elif isinstance(file_data, io.BytesIO):
                # Binary data (PDF, Excel)
                response = HttpResponse(file_data.getvalue(), content_type=content_type)
            else:
                raise ExportError("Invalid file data type")
            
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Error creating download response: {str(e)}")
            raise ExportError(f"Failed to create download response: {str(e)}")
    
    @staticmethod
    def get_content_type(format_type: str) -> str:
        """Get content type for export format"""
        content_types = {
            'csv': 'text/csv',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'pdf': 'application/pdf',
            'html': 'text/html',
            'json': 'application/json'
        }
        
        return content_types.get(format_type.lower(), 'application/octet-stream')
    
    @staticmethod
    def get_file_extension(format_type: str) -> str:
        """Get file extension for export format"""
        extensions = {
            'csv': '.csv',
            'excel': '.xlsx',
            'pdf': '.pdf',
            'html': '.html',
            'json': '.json'
        }
        
        return extensions.get(format_type.lower(), '.txt')